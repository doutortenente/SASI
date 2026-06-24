#!/usr/bin/env python3
"""Converte bundle Files_to_claudecode → .md em staging. Não apaga fontes."""

from __future__ import annotations

import re
import sys
from datetime import date
from pathlib import Path

SRC = Path("/home/dr/Downloads/Files_to_claudecode")
STAGING = Path("/home/dr/dev/sasi/docs/import/claudecode-2026-06-24")
MANIFEST: list[str] = []


def slugify(name: str) -> str:
    base = Path(name).stem
    base = re.sub(r"[^\w\-]+", "-", base, flags=re.UNICODE)
    base = re.sub(r"-+", "-", base).strip("-").lower()
    return base or "document"


def frontmatter(title: str, source: Path, kind: str, notes: str = "") -> str:
    lines = [
        "---",
        f"title: {title}",
        f"source: {source}",
        f"imported: {date.today().isoformat()}",
        f"kind: {kind}",
    ]
    if notes:
        lines.append(f"notes: {notes}")
    lines.extend(["---", ""])
    return "\n".join(lines)


def redact_secrets(text: str) -> str:
    patterns = [
        (r"postgresql://[^\s]+", "postgresql://[REDACTED]"),
        (r"sb_publishable_[A-Za-z0-9_]+", "sb_publishable_[REDACTED]"),
        (r"eyJ[A-Za-z0-9_\-]+\.[A-Za-z0-9_\-]+\.[A-Za-z0-9_\-]+", "eyJ...[REDACTED]"),
        (r"(?i)(senha|password|api[_-]?key)\s*[:=]\s*\S+", r"\1: [REDACTED]"),
        (r"Beiseboro51@", "[REDACTED]"),
    ]
    for pat, repl in patterns:
        text = re.sub(pat, repl, text)
    return text


def write_md(rel: str, title: str, source: Path, body: str, kind: str, notes: str = "") -> Path:
    out = STAGING / rel
    out.parent.mkdir(parents=True, exist_ok=True)
    content = frontmatter(title, source, kind, notes) + body.strip() + "\n"
    if kind in {"handoff", "archive"}:
        content = redact_secrets(content)
    out.write_text(content, encoding="utf-8")
    MANIFEST.append(f"{out.relative_to(STAGING)} ← {source}")
    return out


def convert_docx(path: Path) -> str:
    from docx import Document

    doc = Document(path)
    blocks: list[str] = []
    for p in doc.paragraphs:
        t = p.text.strip()
        if not t:
            continue
        style = (p.style.name or "").lower()
        if "heading 1" in style:
            blocks.append(f"# {t}")
        elif "heading 2" in style:
            blocks.append(f"## {t}")
        elif "heading 3" in style:
            blocks.append(f"## {t}")
        elif t.startswith("#"):
            blocks.append(t)
        else:
            blocks.append(t)
    return "\n\n".join(blocks) if blocks else "_[docx vazio ou só tabelas — revisar manualmente]_"


def convert_pdf(path: Path) -> str:
    from pypdf import PdfReader

    reader = PdfReader(str(path))
    parts: list[str] = []
    for i, page in enumerate(reader.pages, 1):
        text = (page.extract_text() or "").strip()
        if text:
            parts.append(f"## Página {i}\n\n{text}")
    body = "\n\n".join(parts)
    if len(body.strip()) < 80:
        return (
            "_[PDF com pouco texto extraível — provável scan manuscrito. "
            "Use a imagem/PDF original como fixture visual.]_"
        )
    return body


def sheet_to_md_table(rows: list[list]) -> str:
    if not rows:
        return ""
    width = max(len(r) for r in rows)
    norm = [list(r) + [""] * (width - len(r)) for r in rows]
    header = [str(c or "").replace("|", "\\|").replace("\n", " ") for c in norm[0]]
    lines = ["| " + " | ".join(header) + " |", "| " + " | ".join(["---"] * len(header)) + " |"]
    for row in norm[1:]:
        cells = [str(c or "").replace("|", "\\|").replace("\n", " ") for c in row]
        lines.append("| " + " | ".join(cells) + " |")
    return "\n".join(lines)


def convert_xlsx(path: Path) -> str:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    parts: list[str] = []
    for name in wb.sheetnames:
        ws = wb[name]
        rows = [[cell for cell in row] for row in ws.iter_rows(values_only=True)]
        rows = [r for r in rows if any(c not in (None, "") for c in r)]
        if not rows:
            continue
        parts.append(f"## Sheet: {name}\n\n{sheet_to_md_table(rows)}")
    return "\n\n".join(parts) if parts else "_[planilha vazia]_"


def convert_ods(path: Path) -> str:
    from odf.opendocument import load
    from odf.table import Table, TableRow, TableCell
    from odf.text import P

    doc = load(str(path))
    parts: list[str] = []
    for table in doc.spreadsheet.getElementsByType(Table):
        name = table.getAttribute("name") or "Sheet"
        rows: list[list[str]] = []
        for row in table.getElementsByType(TableRow):
            cells: list[str] = []
            for cell in row.getElementsByType(TableCell):
                texts = ["".join(p.childNodes[0].data for p in cell.getElementsByType(P) if p.childNodes)]
                cells.append(" ".join(texts).strip())
            if any(cells):
                rows.append(cells)
        if rows:
            parts.append(f"## Sheet: {name}\n\n{sheet_to_md_table(rows)}")
    return "\n\n".join(parts) if parts else "_[ods vazio]_"


def mirror_existing_md(path: Path, rel_prefix: str) -> None:
    rel = f"{rel_prefix}/{path.name}"
    body = path.read_text(encoding="utf-8", errors="replace")
    if body.startswith("---"):
        write_md(rel, path.stem, path, body.split("---", 2)[-1], "markdown-source")
    else:
        write_md(rel, path.stem, path, body, "markdown-source")


def main() -> int:
    if not SRC.exists():
        print(f"Fonte não encontrada: {SRC}", file=sys.stderr)
        return 1

    STAGING.mkdir(parents=True, exist_ok=True)

    # 1) Espelhar .md / .ts / .json / .sql já legíveis (como referência)
    for ext, folder in [
        (".md", "sources-md"),
        (".ts", "sources-ts"),
        (".json", "sources-json"),
        (".sql", "sources-sql"),
    ]:
        for path in SRC.rglob(f"*{ext}"):
            if ext == ".md":
                mirror_existing_md(path, folder)
            else:
                rel = f"{folder}/{path.relative_to(SRC)}"
                out = STAGING / rel
                out.parent.mkdir(parents=True, exist_ok=True)
                out.write_text(path.read_text(encoding="utf-8", errors="replace"), encoding="utf-8")
                MANIFEST.append(f"{rel} ← {path}")

    # 2) Converter binários
    for path in sorted(SRC.rglob("*")):
        if not path.is_file():
            continue
        ext = path.suffix.lower()
        rel_base = slugify(path.relative_to(SRC).as_posix())

        if ext == ".docx":
            body = convert_docx(path)
            write_md(f"converted-docx/{rel_base}.md", path.stem, path, body, "docx")
        elif ext == ".pdf":
            body = convert_pdf(path)
            notes = "fixture-clinico" if "exemplos" in str(path).lower() else "pdf"
            write_md(f"converted-pdf/{rel_base}.md", path.stem, path, body, "pdf", notes)
        elif ext == ".xlsx":
            body = convert_xlsx(path)
            write_md(f"converted-xlsx/{rel_base}.md", path.stem, path, body, "xlsx")
        elif ext == ".ods":
            body = convert_ods(path)
            write_md(f"converted-ods/{rel_base}.md", path.stem, path, body, "ods")
        elif ext == ".png":
            rel = f"assets/{path.name}"
            out = STAGING / rel
            out.parent.mkdir(parents=True, exist_ok=True)
            out.write_bytes(path.read_bytes())
            MANIFEST.append(f"{rel} ← {path}")

    # 3) Manifest
    (STAGING / "MANIFEST.md").write_text(
        frontmatter("Manifest import claudecode", SRC, "manifest")
        + "\n".join(f"- `{line}`" for line in MANIFEST)
        + "\n",
        encoding="utf-8",
    )

    print(f"Staging: {STAGING}")
    print(f"Arquivos: {len(MANIFEST)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())