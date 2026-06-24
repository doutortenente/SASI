#!/usr/bin/env python3
"""Ingestão: ~/Downloads/Files_to_claudecode → markdown em área de staging."""

from __future__ import annotations

import hashlib
import re
import sys
from datetime import date
from pathlib import Path

BUNDLE_ROOT = Path.home() / "Downloads" / "Files_to_claudecode"
OUT_ROOT = Path(__file__).resolve().parents[1] / "docs" / "import" / "bundle-claudecode"
LOG: list[str] = []


def safe_name(path: Path) -> str:
    stem = re.sub(r"[^a-zA-Z0-9]+", "-", path.stem).strip("-").lower()
    return stem or "file"


def yaml_header(title: str, origin: Path, kind: str, extra: str = "") -> str:
    rows = ["---", f"title: {title}", f"origin: {origin}", f"ingested: {date.today()}", f"kind: {kind}"]
    if extra:
        rows.append(f"note: {extra}")
    rows.append("---\n")
    return "\n".join(rows)


def scrub_credentials(text: str) -> str:
    subs = [
        (r"postgresql://\S+", "postgresql://[REDACTED]"),
        (r"sb_publishable_\w+", "sb_publishable_[REDACTED]"),
        (r"Beiseboro51@", "[REDACTED]"),
        (r"(?i)(senha|password)\s*[:=]\s*\S+", r"\1: [REDACTED]"),
    ]
    for pattern, repl in subs:
        text = re.sub(pattern, repl, text)
    return text


def emit(rel_path: str, title: str, origin: Path, body: str, kind: str, scrub: bool = False, note: str = "") -> None:
    target = OUT_ROOT / rel_path
    target.parent.mkdir(parents=True, exist_ok=True)
    payload = yaml_header(title, origin, kind, note) + (scrub_credentials(body) if scrub else body).strip() + "\n"
    target.write_text(payload, encoding="utf-8")
    LOG.append(f"{rel_path} <= {origin}")


def docx_to_markdown(path: Path) -> str:
    from docx import Document

    chunks: list[str] = []
    for para in Document(path).paragraphs:
        line = para.text.strip()
        if not line:
            continue
        style = (para.style.name or "").lower()
        if "heading 1" in style:
            chunks.append(f"# {line}")
        elif "heading 2" in style or "heading 3" in style:
            chunks.append(f"## {line}")
        else:
            chunks.append(line)
    return "\n\n".join(chunks) if chunks else "_Sem texto extraível no docx._"


def pdf_to_markdown(path: Path) -> str:
    from pypdf import PdfReader

    pages: list[str] = []
    for n, page in enumerate(PdfReader(str(path)).pages, start=1):
        txt = (page.extract_text() or "").strip()
        if txt:
            pages.append(f"### Folha {n}\n\n{txt}")
    joined = "\n\n".join(pages)
    if len(joined) < 60:
        return "_PDF escaneado/manuscrito — pouco texto automático; usar imagem original._"
    return joined


def rows_to_md_table(matrix: list[list]) -> str:
    if not matrix:
        return ""
    cols = max(len(r) for r in matrix)
    padded = [list(r) + [""] * (cols - len(r)) for r in matrix]
    head = [str(c or "").replace("|", "/").replace("\n", " ") for c in padded[0]]
    out = ["| " + " | ".join(head) + " |", "| " + " | ".join(["---"] * len(head)) + " |"]
    for row in padded[1:]:
        out.append("| " + " | ".join(str(c or "").replace("|", "/").replace("\n", " ") for c in row) + " |")
    return "\n".join(out)


def xlsx_to_markdown(path: Path) -> str:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    blocks: list[str] = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rows = [[c for c in row] for row in ws.iter_rows(values_only=True)]
        rows = [r for r in rows if any(c not in (None, "") for c in r)]
        if rows:
            blocks.append(f"## Aba `{sheet}`\n\n{rows_to_md_table(rows)}")
    return "\n\n".join(blocks) if blocks else "_Planilha sem linhas úteis._"


def ods_to_markdown(path: Path) -> str:
    from odf.opendocument import load
    from odf.table import Table, TableRow, TableCell
    from odf.text import P

    doc = load(str(path))
    blocks: list[str] = []
    for table in doc.spreadsheet.getElementsByType(Table):
        name = table.getAttribute("name") or "sheet"
        matrix: list[list[str]] = []
        for row in table.getElementsByType(TableRow):
            cells = []
            for cell in row.getElementsByType(TableCell):
                parts = []
                for p in cell.getElementsByType(P):
                    if p.childNodes:
                        parts.append("".join(getattr(n, "data", str(n)) for n in p.childNodes))
                cells.append(" ".join(parts).strip())
            if any(cells):
                matrix.append(cells)
        if matrix:
            blocks.append(f"## Aba `{name}`\n\n{rows_to_md_table(matrix)}")
    return "\n\n".join(blocks) if blocks else "_ODS vazio._"


def mirror_text_file(path: Path, bucket: str) -> None:
    rel = f"{bucket}/{path.relative_to(BUNDLE_ROOT)}"
    body = path.read_text(encoding="utf-8", errors="replace")
    if body.startswith("---"):
        body = body.split("---", 2)[-1]
    scrub = "_HANDOFF" in path.name or "README_SYNC" in path.name
    emit(str(rel), path.stem, path, body, "source-text", scrub=scrub)


def main() -> int:
    if not BUNDLE_ROOT.is_dir():
        print(f"Bundle ausente: {BUNDLE_ROOT}", file=sys.stderr)
        return 1

    OUT_ROOT.mkdir(parents=True, exist_ok=True)

    for path in sorted(BUNDLE_ROOT.rglob("*")):
        if not path.is_file():
            continue
        ext = path.suffix.lower()
        tag = safe_name(path.relative_to(BUNDLE_ROOT))

        if ext == ".md":
            mirror_text_file(path, "raw-markdown")
        elif ext == ".ts":
            rel = f"raw-typescript/{path.relative_to(BUNDLE_ROOT)}"
            (OUT_ROOT / rel).parent.mkdir(parents=True, exist_ok=True)
            (OUT_ROOT / rel).write_text(path.read_text(encoding="utf-8"), encoding="utf-8")
            LOG.append(f"{rel} <= {path}")
        elif ext in {".json", ".sql"}:
            rel = f"raw-data/{path.relative_to(BUNDLE_ROOT)}"
            (OUT_ROOT / rel).parent.mkdir(parents=True, exist_ok=True)
            (OUT_ROOT / rel).write_text(path.read_text(encoding="utf-8"), encoding="utf-8")
            LOG.append(f"{rel} <= {path}")
        elif ext == ".docx":
            emit(f"from-docx/{tag}.md", path.stem, path, docx_to_markdown(path), "docx")
        elif ext == ".pdf":
            note = "clinical-fixture" if "exemplos" in str(path).lower() else "pdf"
            emit(f"from-pdf/{tag}.md", path.stem, path, pdf_to_markdown(path), "pdf", note=note)
        elif ext == ".xlsx":
            emit(f"from-xlsx/{tag}.md", path.stem, path, xlsx_to_markdown(path), "xlsx")
        elif ext == ".ods":
            emit(f"from-ods/{tag}.md", path.stem, path, ods_to_markdown(path), "ods")
        elif ext == ".png":
            rel = f"binaries/{path.name}"
            (OUT_ROOT / rel).parent.mkdir(parents=True, exist_ok=True)
            (OUT_ROOT / rel).write_bytes(path.read_bytes())
            LOG.append(f"{rel} <= {path}")

    emit(
        "INGEST_LOG.md",
        "Ingest log",
        BUNDLE_ROOT,
        "\n".join(f"- `{line}`" for line in LOG),
        "log",
    )
    print(f"OK {OUT_ROOT} ({len(LOG)} itens)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())